"""
Excel file reader for syncing variable values from Excel cells.
"""

import os
import uuid
from typing import Optional
from openpyxl import load_workbook


# GUID property name for tracking Excel files
TANSU_GUID_PROPERTY = "TansuGUID"


def get_excel_guid(file_path: str) -> Optional[str]:
    """
    Read TansuGUID from Excel custom document properties.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        GUID string if found, None otherwise
    """
    if not os.path.exists(file_path):
        return None

    try:
        wb = load_workbook(file_path, read_only=False)

        # Check custom document properties
        if wb.custom_doc_props and TANSU_GUID_PROPERTY in wb.custom_doc_props:
            guid = wb.custom_doc_props[TANSU_GUID_PROPERTY].value
            wb.close()
            return str(guid) if guid else None

        wb.close()
        return None
    except Exception:
        return None


def set_excel_guid(file_path: str, guid: str = None) -> Optional[str]:
    """
    Write TansuGUID to Excel custom document properties.
    If guid is None, generates a new UUID.

    Args:
        file_path: Path to the .xlsx file
        guid: Optional GUID to set, generates new one if None

    Returns:
        The GUID that was set, or None if failed
    """
    if not os.path.exists(file_path):
        return None

    if guid is None:
        guid = str(uuid.uuid4())

    try:
        from openpyxl.packaging.custom import CustomPropertyList, StringProperty

        wb = load_workbook(file_path, read_only=False)

        # Create or get custom properties
        if wb.custom_doc_props is None:
            wb.custom_doc_props = CustomPropertyList()

        # Remove existing property if present
        existing = [p for p in wb.custom_doc_props if p.name == TANSU_GUID_PROPERTY]
        for p in existing:
            wb.custom_doc_props.remove(p)

        # Add new property
        wb.custom_doc_props.append(StringProperty(name=TANSU_GUID_PROPERTY, value=guid))

        wb.save(file_path)
        wb.close()
        return guid
    except Exception as e:
        # File might be read-only or open in another app
        return None


def get_or_create_excel_guid(file_path: str) -> Optional[str]:
    """
    Get existing GUID from Excel file, or create and store a new one.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        The GUID (existing or newly created), or None if failed
    """
    # Try to read existing GUID
    guid = get_excel_guid(file_path)
    if guid:
        return guid

    # Create new GUID
    return set_excel_guid(file_path)


def read_cell_value(file_path: str, sheet_name: str, cell_ref: str) -> Optional[str]:
    """
    Read a single cell value from an Excel file.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        cell_ref: Cell reference (e.g., 'A1', 'B5', 'C10')

    Returns:
        Cell value as string, or None if not found
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {os.path.basename(file_path)}")

    ws = wb[sheet_name]
    value = ws[cell_ref].value
    wb.close()

    if value is None:
        return ""

    # Convert to string, handling numbers nicely
    if isinstance(value, float):
        # Remove trailing zeros for cleaner display
        if value == int(value):
            return str(int(value))
        return str(value)

    return str(value)


def read_sheet_preview(file_path: str, sheet_name: str, max_rows: int = 20, max_cols: int = 10) -> list[list[str]]:
    """
    Read a preview of the sheet as a 2D list of cell values.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        max_rows: Maximum rows to read
        max_cols: Maximum columns to read

    Returns:
        2D list of cell values (strings)
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    data = []
    for row_idx in range(1, max_rows + 1):
        row_data = []
        for col_idx in range(1, max_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                row_data.append("")
            elif isinstance(value, float):
                if value == int(value):
                    row_data.append(str(int(value)))
                else:
                    row_data.append(str(value))
            else:
                row_data.append(str(value))
        data.append(row_data)

    wb.close()
    return data


def get_sheet_names(file_path: str) -> list[str]:
    """
    Get list of sheet names in an Excel file.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        List of sheet names
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def sync_variables_from_excel(variables: list[dict]) -> dict[int, tuple[str, str]]:
    """
    Sync multiple variables from their linked Excel cells.

    Args:
        variables: List of variable dicts with excel_file, excel_sheet, excel_cell

    Returns:
        Dict of var_id -> (old_value, new_value) for variables that changed
    """
    changes = {}

    for var in variables:
        var_id = var['id']
        file_path = var.get('excel_file')
        sheet_name = var.get('excel_sheet')
        cell_ref = var.get('excel_cell')

        if not all([file_path, sheet_name, cell_ref]):
            continue

        try:
            new_value = read_cell_value(file_path, sheet_name, cell_ref)
            old_value = var.get('value', '')

            if new_value != old_value:
                changes[var_id] = (old_value, new_value)
        except Exception:
            # Skip variables that can't be read
            pass

    return changes


def read_range_as_variables(file_path: str, sheet_name: str, start_cell: str) -> list[dict]:
    """
    Read a range of cells as variables (Name, Value, Unit columns).
    Reads from start_cell down until it hits an empty Name cell.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        start_cell: Top-left cell of the range (e.g., 'A1', 'B5')

    Returns:
        List of dicts with 'name', 'value', 'unit' keys
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {os.path.basename(file_path)}")

    ws = wb[sheet_name]

    # Parse start cell to get column and row
    import re
    match = re.match(r'([A-Za-z]+)(\d+)', start_cell.upper())
    if not match:
        wb.close()
        raise ValueError(f"Invalid cell reference: {start_cell}")

    start_col = match.group(1)
    start_row = int(match.group(2))

    # Convert column letter to number (A=1, B=2, etc.)
    def col_to_num(col_str):
        result = 0
        for char in col_str:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    def num_to_col(num):
        result = ""
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    col_num = col_to_num(start_col)
    name_col = start_col
    value_col = num_to_col(col_num + 1)
    unit_col = num_to_col(col_num + 2)

    variables = []
    row = start_row
    empty_rows = 0
    max_empty_rows = 5  # Skip up to 5 empty rows at the start to find data

    while True:
        # Read name cell
        name_cell = ws[f"{name_col}{row}"]
        name_value = name_cell.value

        # Handle empty rows
        if name_value is None or str(name_value).strip() == "":
            # If we haven't found any data yet, skip empty rows
            if not variables and empty_rows < max_empty_rows:
                empty_rows += 1
                row += 1
                continue
            # If we already have data, stop at first empty row
            break

        name = str(name_value).strip().replace(' ', '_')

        # Read value cell
        value_cell = ws[f"{value_col}{row}"]
        value = value_cell.value
        if value is None:
            value = ""
        elif isinstance(value, float):
            if value == int(value):
                value = str(int(value))
            else:
                value = str(value)
        else:
            value = str(value).strip()

        # Read unit cell (optional)
        unit_cell = ws[f"{unit_col}{row}"]
        unit = unit_cell.value
        if unit is None:
            unit = ""
        else:
            unit = str(unit).strip()

        variables.append({
            'name': name,
            'value': value,
            'unit': unit,
            'row': row  # Store row for reference
        })

        row += 1

        # Safety limit
        if row > start_row + 1000:
            break

    wb.close()
    return variables


def validate_excel_range(file_path: str, sheet_name: str, start_cell: str) -> tuple[bool, str, list[dict]]:
    """
    Validate an Excel range and return preview of variables.

    Returns:
        Tuple of (is_valid, message, variables_list)
    """
    if not file_path:
        return False, "No file path specified", []

    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}", []

    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        return False, "File must be .xlsx or .xlsm format", []

    try:
        sheets = get_sheet_names(file_path)
        if sheet_name not in sheets:
            return False, f"Sheet '{sheet_name}' not found. Available: {', '.join(sheets)}", []

        variables = read_range_as_variables(file_path, sheet_name, start_cell)

        if not variables:
            return False, "No variables found starting at that cell", []

        return True, f"Found {len(variables)} variable(s)", variables
    except Exception as e:
        return False, f"Error: {e}", []


def validate_excel_link(file_path: str, sheet_name: str, cell_ref: str) -> tuple[bool, str]:
    """
    Validate that an Excel link is valid and readable.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        cell_ref: Cell reference

    Returns:
        Tuple of (is_valid, message_or_value)
    """
    if not file_path:
        return False, "No file path specified"

    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"

    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        return False, "File must be .xlsx or .xlsm format"

    try:
        sheets = get_sheet_names(file_path)
        if sheet_name not in sheets:
            return False, f"Sheet '{sheet_name}' not found. Available: {', '.join(sheets)}"

        value = read_cell_value(file_path, sheet_name, cell_ref)
        return True, f"Current value: {value if value else '(empty)'}"
    except Exception as e:
        return False, f"Error: {e}"
